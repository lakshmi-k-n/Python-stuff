import csv
import openpyxl
import os


for file in os.listdir('.'):
    if file.endswith('.xlsx'):
        wb = openpyxl.load_workbook(file)

        for sheet1 in wb.get_sheet1s():
            sheet = wb.get_sheet_by_name(sheet1)
            namee = file[: -5]


            csvfile = open(namee + '_' + sheet1 + '.csv', 'w', newline='')
            csvwriter = csv.writer(csvfile)

            for rowcount in range(1, sheet.max_row + 1):
                rdata = []
                for colcount in range(1, sheet.max_column + 1):
                    
                   rdata.append(sheet.cell(row=rowcount, column=colcount).value)

                for row in row_data:
                    csv_writer.writerow(row)
csv_file.close()
